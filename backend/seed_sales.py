"""Generate seed Sales Dashboard Excel with 3 sheets"""
import openpyxl
import random
from pathlib import Path

DATA_DIR = Path(__file__).parent / 'data'
DATA_DIR.mkdir(exist_ok=True)

PROJECTS = [
    {"name": "Smartworld Sky Arc", "short": "Sky Arc", "code": "1070"},
    {"name": "Smartworld The Edition", "short": "The Edition", "code": "1072"},
    {"name": "Smartworld Trump Residences", "short": "Trump Residences", "code": "1073"},
]

TOWERS = ["Tower A", "Tower B", "Tower C", "Tower D"]
UNIT_TYPES = ["2BHK", "3BHK", "4BHK", "Penthouse"]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul"]
LOAN_STATUSES = ["Approved", "Pending", "Not Applied", "Rejected"]
CHANNEL_TYPES = ["Channel Partner", "Direct"]
BANKS = ["SBI", "HDFC", "ICICI", "Axis", "PNB", "Kotak", ""]
UNIT_STATUSES = ["Available", "Booked", "Allotted"]
PAYMENT_MODES = ["ONLINE", "CREDIT", "CHEQUE", "TDS CHALLAN", "REBATE"]
RECEIPT_STATUSES = ["CLEARED", "PAYMENT", "ADJUSTMENT", "BOUNCE", "CANCELLED"]
MILESTONES_BILLED = [
    "Booking Amount", "On Allotment", "On Booking",
    "On Completion of Ground Floor", "On Completion of Excavation Work",
    "On or Before 01st Aug 2025", "On or Before 01st July 2025 (On Execution/Signing of Agreement for Sale)",
    "On or Before 01st June 2025 (On Execution/Signing of Agreement for Sale)",
    "On or Before 01st March 2025", "On or Before 01st November 2024 (On Execution)"
]
MILESTONES_UNBILLED = [
    "On Notice of Offer of Possession", "On Application of Occupation Certificate",
    "On Offer of Possession", "On Laying of 40th Floor",
    "Within 18 months from Allotment Or On casting of 5th Floor slab",
    "Within 25 months from Allotment Or On completion of 20th floor",
    "Within 30 months from Allotment Or On completion of 30th floor",
    "Within 36 months from Allotment Or On completion of 40th floor",
    "Within 42 months from Allotment"
]
AGEING_BUCKETS = ["1-30 Days", "31-90 Days", "91-180 Days", "181+ Days"]


def gen_sales():
    """Generate ~2200 sales records matching dashboard values"""
    rows = []
    # Target: Total 2202 units, 533 available, 1669 allotted (1667 booked)
    # Sky Arc: 4649 Cr sales, Edition: 3425 Cr, Trump: 2348 Cr
    targets = {
        "Sky Arc": {"sales": 4649, "demand": 1133, "received": 1127, "units": 800},
        "The Edition": {"sales": 3425, "demand": 1170, "received": 1014, "units": 750},
        "Trump Residences": {"sales": 2348, "demand": 667, "received": 625, "units": 652},
    }
    # Monthly sales targets
    monthly_sales = {"Jan": 326, "Feb": 337, "Mar": 243, "Apr": 196, "May": 2132, "Jun": 274, "Jul": 104}

    unit_counter = 21100000
    for proj in PROJECTS:
        t = targets[proj["short"]]
        n_units = t["units"]
        sales_per_unit = t["sales"] / n_units
        demand_per_unit = t["demand"] / n_units
        received_per_unit = t["received"] / n_units

        for i in range(n_units):
            unit_counter += 1
            month = random.choice(MONTHS)
            status = random.choices(UNIT_STATUSES, weights=[24, 60, 16], k=1)[0]
            channel = random.choices(CHANNEL_TYPES, weights=[95, 5], k=1)[0]
            loan = random.choice(LOAN_STATUSES)
            tower = random.choice(TOWERS)
            utype = random.choice(UNIT_TYPES)
            bank = random.choice(BANKS) if loan == "Approved" else ""

            sale_val = round(sales_per_unit * random.uniform(0.6, 1.4), 2)
            dem_val = round(demand_per_unit * random.uniform(0.5, 1.5), 2)
            rec_val = round(min(dem_val, dem_val * random.uniform(0.85, 1.0)), 2)
            outstanding = round(dem_val - rec_val, 2)
            credit_debit = round(random.uniform(-0.1, 0.1), 2) if random.random() < 0.05 else 0

            builtup = round(random.uniform(800, 4500), 0)
            carpet = round(builtup * random.uniform(0.5, 0.6), 0)

            rows.append([
                "Smartworld", proj["name"], proj["short"], proj["code"],
                tower, utype, f"U-{unit_counter}", status, loan, channel,
                month, 2025, str(unit_counter), bank,
                sale_val, dem_val, rec_val, outstanding, credit_debit,
                builtup, carpet
            ])
    return rows


def gen_payments():
    """Generate ~17000 payment records"""
    rows = []
    # Target: Online 6472, Credit 6372, Cheque 2245, TDS Challan 1936, Rebate 28
    mode_targets = {
        "ONLINE": {"count": 6472, "collected": 1778.26, "bounce": 3.02},
        "CREDIT": {"count": 6372, "collected": 808.02, "bounce": 24.0},
        "CHEQUE": {"count": 2245, "collected": 292.52, "bounce": 61.38},
        "TDS CHALLAN": {"count": 1936, "collected": 12.25, "bounce": 0.5},
        "REBATE": {"count": 28, "collected": 1.09, "bounce": 0.01},
    }

    for mode, t in mode_targets.items():
        amt_per = t["collected"] / t["count"] if t["count"] > 0 else 0
        bounce_count = int(t["count"] * 0.03)
        cancel_count = int(t["count"] * 0.01)
        cleared_count = t["count"] - bounce_count - cancel_count

        for i in range(t["count"]):
            proj = random.choice(PROJECTS)
            month = random.choice(MONTHS)
            if i < cleared_count:
                status = "CLEARED"
                amt = round(amt_per * random.uniform(0.3, 1.7), 2)
                bounce_amt = 0
            elif i < cleared_count + bounce_count:
                status = "BOUNCE"
                amt = 0
                bounce_amt = round(t["bounce"] / bounce_count * random.uniform(0.5, 1.5), 2) if bounce_count > 0 else 0
            else:
                status = "CANCELLED"
                amt = 0
                bounce_amt = round(random.uniform(0.01, 0.5), 2)

            is_bank = random.random() < 0.08
            rows.append([
                "Smartworld", proj["name"], proj["short"],
                random.choice(TOWERS), random.choice(UNIT_TYPES),
                f"U-{random.randint(21100001, 21102200)}",
                random.choice(LOAN_STATUSES), month, 2025,
                mode, status, round(amt, 2), round(bounce_amt, 2),
                "Bank" if is_bank else "Self"
            ])
    return rows


def gen_ageing():
    """Generate ageing data"""
    rows = []
    # Billed milestones
    billed_data = [
        ("Booking Amount", "91-180 Days", 19, 9.69),
        ("Booking Amount", "181+ Days", 540, 57.56),
        ("On Allotment", "181+ Days", 210, 92.92),
        ("On Booking", "31-90 Days", 4, 1.05),
        ("On Booking", "181+ Days", 22, 4.13),
        ("On Completion of Ground Floor", "31-90 Days", 2, 0.60),
        ("On Completion of Excavation Work", "91-180 Days", 12, 24.75),
        ("On or Before 01st Aug 2025", "91-180 Days", 1, 0.44),
        ("On or Before 01st July 2025 (On Execution/Signing of Agreement for Sale)", "181+ Days", 1, 1.32),
        ("On or Before 01st June 2025 (On Execution/Signing of Agreement for Sale)", "181+ Days", 1, 0.46),
        ("On or Before 01st March 2025", "181+ Days", 2, 2.28),
        ("On or Before 01st November 2024 (On Execution)", "181+ Days", 1, 0.95),
    ]
    for milestone, bucket, count, total_amt in billed_data:
        per_amt = total_amt / count if count > 0 else 0
        for i in range(count):
            proj = random.choice(PROJECTS)
            rows.append([
                "Smartworld", proj["name"], proj["short"],
                random.choice(UNIT_TYPES), f"U-{random.randint(21100001, 21102200)}",
                random.choice(MONTHS), 2025, bucket,
                milestone, "Billed", 1,
                round(per_amt * random.uniform(0.7, 1.3), 2),
                round(per_amt * random.uniform(2, 4), 2),
                round(per_amt * random.uniform(1, 2.5), 2),
            ])

    # Unbilled milestones
    unbilled_data = [
        ("On Notice of Offer of Possession", 3800, 1268),
        ("On Application of Occupation Certificate", 1500, 958),
        ("On Offer of Possession", 1200, 605),
        ("On Laying of 40th Floor", 800, 416),
        ("Within 18 months from Allotment Or On casting of 5th Floor slab", 800, 416),
        ("Within 25 months from Allotment Or On completion of 20th floor", 800, 416),
        ("Within 30 months from Allotment Or On completion of 30th floor", 800, 403),
        ("Within 36 months from Allotment Or On completion of 40th floor", 800, 403),
        ("Within 42 months from Allotment", 500, 280),
    ]
    for milestone, count, total_amt in unbilled_data:
        sample = min(count, 200)  # Limit rows
        per_amt = total_amt / count if count > 0 else 0
        for i in range(sample):
            proj = random.choice(PROJECTS)
            rows.append([
                "Smartworld", proj["name"], proj["short"],
                random.choice(UNIT_TYPES), f"U-{random.randint(21100001, 21102200)}",
                random.choice(MONTHS), 2025, "",
                milestone, "Unbilled", int(count / sample),
                round(per_amt * (count / sample) * random.uniform(0.7, 1.3), 2),
                round(per_amt * (count / sample) * random.uniform(2, 4), 2),
                round(per_amt * (count / sample) * random.uniform(1, 2.5), 2),
            ])
    return rows


def generate_sales_excel():
    filepath = DATA_DIR / 'Sales.xlsx'
    wb = openpyxl.Workbook()

    # Sheet 1: Sales Data
    ws = wb.active
    ws.title = 'Sales_Data'
    ws.append([
        'Company Name', 'Project Name', 'Project Short Name', 'Project Code',
        'Tower', 'Unit Type', 'Unit No', 'Unit Status', 'Loan Status', 'Channel Type',
        'Month', 'Year', 'Sales Order Number', 'Bank Name',
        'Total Sales (Cr)', 'Demand (Cr)', 'Received (Cr)', 'Outstanding (Cr)', 'Credit/Debit (Cr)',
        'Built-up Area (Sqft)', 'Carpet Area (Sqft)'
    ])
    for row in gen_sales():
        ws.append(row)

    # Sheet 2: Payment Data
    ws2 = wb.create_sheet('Payment_Data')
    ws2.append([
        'Company Name', 'Project Name', 'Project Short Name',
        'Tower', 'Unit Type', 'Unit No', 'Loan Status', 'Month', 'Year',
        'Mode of Payment', 'Receipt Status', 'Collected Amount (Cr)', 'Cancelled/Bounce Amount (Cr)',
        'Funding Type'
    ])
    for row in gen_payments():
        ws2.append(row)

    # Sheet 3: Ageing Data
    ws3 = wb.create_sheet('Ageing_Data')
    ws3.append([
        'Company Name', 'Project Name', 'Project Short Name',
        'Unit Type', 'Unit No', 'Month', 'Year', 'Ageing Bucket',
        'Milestone', 'Billed Status', 'Installment Count', 'Installment Amount (Cr)',
        'Demand Amount (Cr)', 'Received Amount (Cr)'
    ])
    for row in gen_ageing():
        ws3.append(row)

    wb.save(filepath)
    print(f"Sales Excel saved to {filepath}")
    return str(filepath)


if __name__ == '__main__':
    generate_sales_excel()
