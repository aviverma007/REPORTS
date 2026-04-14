"""Sales Dashboard API routes"""
from fastapi import APIRouter, UploadFile, File, HTTPException
from typing import Optional
from pathlib import Path
import openpyxl
import shutil

DATA_DIR = Path(__file__).parent / 'data'
SALES_FILE = DATA_DIR / 'Sales.xlsx'

sales_router = APIRouter(prefix="/api/sales")


def ensure_sales_data():
    if not SALES_FILE.exists():
        from seed_sales import generate_sales_excel
        generate_sales_excel()


def read_sheet(sheet_name):
    ensure_sales_data()
    wb = openpyxl.load_workbook(SALES_FILE, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                d[headers[i]] = val
        rows.append(d)
    wb.close()
    return rows


def apply_filters(rows, params):
    filtered = rows
    for key, val in params.items():
        if val:
            # Map filter keys to column names
            col_map = {
                'company': 'Company Name', 'project': 'Project Short Name',
                'tower': 'Tower', 'unit_type': 'Unit Type', 'unit_no': 'Unit No',
                'loan_status': 'Loan Status', 'month': 'Month', 'year': 'Year',
                'ageing_bucket': 'Ageing Bucket'
            }
            col = col_map.get(key)
            if col:
                filtered = [r for r in filtered if str(r.get(col, '')) == str(val)]
    return filtered


@sales_router.get("/filters")
async def get_sales_filters():
    sales = read_sheet('Sales_Data')
    companies = sorted(set(str(r.get('Company Name', '')) for r in sales if r.get('Company Name')))
    projects = sorted(set(str(r.get('Project Short Name', '')) for r in sales if r.get('Project Short Name')))
    towers = sorted(set(str(r.get('Tower', '')) for r in sales if r.get('Tower')))
    unit_types = sorted(set(str(r.get('Unit Type', '')) for r in sales if r.get('Unit Type')))
    loan_statuses = sorted(set(str(r.get('Loan Status', '')) for r in sales if r.get('Loan Status')))
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    avail_months = [m for m in months if m in set(str(r.get('Month', '')) for r in sales)]
    years = sorted(set(str(int(r.get('Year'))) for r in sales if r.get('Year')))

    return {
        "companies": companies, "projects": projects, "towers": towers,
        "unit_types": unit_types, "loan_statuses": loan_statuses,
        "months": avail_months, "years": years
    }


@sales_router.get("/data")
async def get_sales_data(
    company: Optional[str] = None, project: Optional[str] = None,
    tower: Optional[str] = None, unit_type: Optional[str] = None,
    unit_no: Optional[str] = None, loan_status: Optional[str] = None,
    month: Optional[str] = None, year: Optional[str] = None
):
    params = {
        'company': company, 'project': project, 'tower': tower,
        'unit_type': unit_type, 'unit_no': unit_no, 'loan_status': loan_status,
        'month': month, 'year': year
    }

    sales = apply_filters(read_sheet('Sales_Data'), params)

    # KPIs
    total_sales = sum(r.get('Total Sales (Cr)', 0) or 0 for r in sales)
    demand = sum(r.get('Demand (Cr)', 0) or 0 for r in sales)
    received = sum(r.get('Received (Cr)', 0) or 0 for r in sales)
    outstanding = sum(r.get('Outstanding (Cr)', 0) or 0 for r in sales)
    credit_debit = sum(r.get('Credit/Debit (Cr)', 0) or 0 for r in sales)

    # Unit counts by status
    status_counts = {}
    for r in sales:
        s = r.get('Unit Status', 'Unknown')
        status_counts[s] = status_counts.get(s, 0) + 1
    total_units = len(sales)
    available = status_counts.get('Available', 0)
    booked = status_counts.get('Booked', 0)
    allotted = status_counts.get('Allotted', 0)

    # Channel type (booked units only)
    booked_rows = [r for r in sales if r.get('Unit Status') in ('Booked', 'Allotted')]
    channel_counts = {}
    for r in booked_rows:
        ch = r.get('Channel Type', 'Unknown')
        channel_counts[ch] = channel_counts.get(ch, 0) + 1

    # Built-up and Carpet area
    total_builtup = sum(r.get('Built-up Area (Sqft)', 0) or 0 for r in sales)
    booked_builtup = sum(r.get('Built-up Area (Sqft)', 0) or 0 for r in booked_rows)
    total_carpet = sum(r.get('Carpet Area (Sqft)', 0) or 0 for r in sales)
    booked_carpet = sum(r.get('Carpet Area (Sqft)', 0) or 0 for r in booked_rows)

    # Project-wise sales
    proj_sales = {}
    for r in sales:
        p = r.get('Project Short Name', '')
        if p not in proj_sales:
            proj_sales[p] = {"project": p, "sales": 0, "demand": 0, "received": 0, "outstanding": 0}
        proj_sales[p]["sales"] += r.get('Total Sales (Cr)', 0) or 0
        proj_sales[p]["demand"] += r.get('Demand (Cr)', 0) or 0
        proj_sales[p]["received"] += r.get('Received (Cr)', 0) or 0
        proj_sales[p]["outstanding"] += r.get('Outstanding (Cr)', 0) or 0
    proj_list = sorted(proj_sales.values(), key=lambda x: x['sales'], reverse=True)

    # Month-wise sales
    month_order = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6, "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}
    month_sales = {}
    for r in sales:
        m = r.get('Month', '')
        if m not in month_sales:
            month_sales[m] = {"month": m, "sales": 0, "demand": 0, "received": 0}
        month_sales[m]["sales"] += r.get('Total Sales (Cr)', 0) or 0
        month_sales[m]["demand"] += r.get('Demand (Cr)', 0) or 0
        month_sales[m]["received"] += r.get('Received (Cr)', 0) or 0
    month_list = sorted(month_sales.values(), key=lambda x: month_order.get(x['month'], 99))

    return {
        "kpi": {
            "total_sales": round(total_sales, 2), "demand": round(demand, 2),
            "received": round(received, 2), "outstanding": round(outstanding, 2),
            "credit_debit": round(credit_debit, 2)
        },
        "units": {
            "total": total_units, "available": available,
            "booked": booked, "allotted": allotted
        },
        "channel_types": [{"name": k, "value": v} for k, v in channel_counts.items()],
        "area": {
            "total_builtup": round(total_builtup), "booked_builtup": round(booked_builtup),
            "leftover_builtup": round(total_builtup - booked_builtup),
            "total_carpet": round(total_carpet), "booked_carpet": round(booked_carpet),
            "leftover_carpet": round(total_carpet - booked_carpet)
        },
        "project_sales": proj_list,
        "month_sales": month_list,
        "row_count": len(sales)
    }


@sales_router.get("/payment")
async def get_payment_data(
    company: Optional[str] = None, project: Optional[str] = None,
    tower: Optional[str] = None, unit_type: Optional[str] = None,
    loan_status: Optional[str] = None, month: Optional[str] = None,
    year: Optional[str] = None
):
    params = {
        'company': company, 'project': project, 'tower': tower,
        'unit_type': unit_type, 'loan_status': loan_status,
        'month': month, 'year': year
    }
    payments = apply_filters(read_sheet('Payment_Data'), params)

    total_collected = sum(r.get('Collected Amount (Cr)', 0) or 0 for r in payments)
    total_bounce = sum(r.get('Cancelled/Bounce Amount (Cr)', 0) or 0 for r in payments)

    self_funded = sum(r.get('Collected Amount (Cr)', 0) or 0 for r in payments if r.get('Funding Type') == 'Self')
    bank_funded = sum(r.get('Collected Amount (Cr)', 0) or 0 for r in payments if r.get('Funding Type') == 'Bank')

    # By mode
    mode_data = {}
    for r in payments:
        m = r.get('Mode of Payment', '')
        if m not in mode_data:
            mode_data[m] = {"mode": m, "count": 0, "collected": 0, "bounce": 0}
        mode_data[m]["count"] += 1
        mode_data[m]["collected"] += r.get('Collected Amount (Cr)', 0) or 0
        mode_data[m]["bounce"] += r.get('Cancelled/Bounce Amount (Cr)', 0) or 0
    mode_list = sorted(mode_data.values(), key=lambda x: x['count'], reverse=True)

    # By mode + receipt status for stacked bar
    mode_status = {}
    for r in payments:
        m = r.get('Mode of Payment', '')
        s = r.get('Receipt Status', '')
        key = f"{m}_{s}"
        if key not in mode_status:
            mode_status[key] = {"mode": m, "status": s, "count": 0}
        mode_status[key]["count"] += 1

    return {
        "kpi": {
            "total_collected": round(total_collected, 2),
            "self_funded": round(self_funded, 2),
            "bank_funded": round(bank_funded, 2),
            "total_bounce": round(total_bounce, 2)
        },
        "total_payments": len(payments),
        "by_mode": mode_list,
        "by_mode_status": list(mode_status.values())
    }


@sales_router.get("/ageing")
async def get_ageing_data(
    company: Optional[str] = None, project: Optional[str] = None,
    unit_type: Optional[str] = None, month: Optional[str] = None,
    year: Optional[str] = None, ageing_bucket: Optional[str] = None
):
    params = {
        'company': company, 'project': project, 'unit_type': unit_type,
        'month': month, 'year': year, 'ageing_bucket': ageing_bucket
    }
    ageing = apply_filters(read_sheet('Ageing_Data'), params)

    demand_total = sum(r.get('Demand Amount (Cr)', 0) or 0 for r in ageing)
    received_total = sum(r.get('Received Amount (Cr)', 0) or 0 for r in ageing)
    outstanding_total = round(demand_total - received_total, 2)

    billed = [r for r in ageing if r.get('Billed Status') == 'Billed']
    unbilled = [r for r in ageing if r.get('Billed Status') == 'Unbilled']

    # Ageing bucket aggregation (billed only)
    bucket_data = {}
    for r in billed:
        b = r.get('Ageing Bucket', '')
        if b not in bucket_data:
            bucket_data[b] = {"bucket": b, "count": 0, "amount": 0}
        bucket_data[b]["count"] += r.get('Installment Count', 0) or 0
        bucket_data[b]["amount"] += r.get('Installment Amount (Cr)', 0) or 0
    bucket_order = {"1-30 Days": 1, "31-90 Days": 2, "91-180 Days": 3, "181+ Days": 4}
    bucket_list = sorted(bucket_data.values(), key=lambda x: bucket_order.get(x['bucket'], 99))

    # Billed milestones table
    milestone_billed = {}
    for r in billed:
        m = r.get('Milestone', '')
        b = r.get('Ageing Bucket', '')
        key = f"{m}_{b}"
        if key not in milestone_billed:
            milestone_billed[key] = {"milestone": m, "ageing_bucket": b, "count": 0, "amount": 0}
        milestone_billed[key]["count"] += r.get('Installment Count', 0) or 0
        milestone_billed[key]["amount"] += r.get('Installment Amount (Cr)', 0) or 0
    billed_table = sorted(milestone_billed.values(), key=lambda x: x['amount'], reverse=True)

    # Unbilled milestones
    milestone_unbilled = {}
    for r in unbilled:
        m = r.get('Milestone', '')
        if m not in milestone_unbilled:
            milestone_unbilled[m] = {"milestone": m, "count": 0, "amount": 0}
        milestone_unbilled[m]["count"] += r.get('Installment Count', 0) or 0
        milestone_unbilled[m]["amount"] += r.get('Installment Amount (Cr)', 0) or 0
    unbilled_list = sorted(milestone_unbilled.values(), key=lambda x: x['count'], reverse=True)

    return {
        "kpi": {
            "demand": round(demand_total, 2),
            "received": round(received_total, 2),
            "outstanding": round(outstanding_total, 2)
        },
        "ageing_buckets": bucket_list,
        "billed_milestones": billed_table,
        "unbilled_milestones": unbilled_list,
        "billed_total_count": sum(r.get('Installment Count', 0) or 0 for r in billed),
        "billed_total_amount": round(sum(r.get('Installment Amount (Cr)', 0) or 0 for r in billed), 2)
    }


@sales_router.post("/upload")
async def upload_sales_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, .csv files are accepted")
    if SALES_FILE.exists():
        shutil.copy2(SALES_FILE, DATA_DIR / 'Sales_backup.xlsx')
    contents = await file.read()
    with open(SALES_FILE, 'wb') as f:
        f.write(contents)
    try:
        wb = openpyxl.load_workbook(SALES_FILE, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return {"message": "File uploaded successfully", "sheets": sheets}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid file: {str(e)}")


@sales_router.get("/download")
async def download_sales():
    from fastapi.responses import FileResponse
    ensure_sales_data()
    return FileResponse(str(SALES_FILE), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename='Sales.xlsx')
