"""Case Management Dashboard API routes"""
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from typing import Optional
from pathlib import Path
import openpyxl
import shutil

DATA_DIR = Path(__file__).parent / 'data'
CASE_FILE = DATA_DIR / 'CaseManagement.xlsx'

case_router = APIRouter(prefix="/api/cases")

# Cache for loaded data
_cache = {"headers": None, "rows": None, "mtime": 0}


def load_case_data():
    if not CASE_FILE.exists():
        return [], []
    mtime = CASE_FILE.stat().st_mtime
    if _cache["rows"] is not None and _cache["mtime"] == mtime:
        return _cache["headers"], _cache["rows"]
    wb = openpyxl.load_workbook(CASE_FILE, read_only=True, data_only=True)
    ws = wb['Sheet1']
    headers = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h) if h else f'col_{j}' for j, h in enumerate(row)]
            continue
        d = {}
        for j, val in enumerate(row):
            if j < len(headers):
                d[headers[j]] = val
        rows.append(d)
    wb.close()
    _cache["headers"] = headers
    _cache["rows"] = rows
    _cache["mtime"] = mtime
    return headers, rows


def safe_str(val):
    if val is None:
        return ''
    return str(val).strip()


def apply_case_filters(rows, params):
    filtered = rows
    col_map = {
        'case_type': 'Case Type',
        'status': 'Status',
        'case_origin': 'Case Origin',
        'area': 'Area',
        'sub_area': 'Sub Area',
        'case_owner': 'Case Owner',
        'hod': 'HOD 1',
        'team_leader': 'Team Leader',
        'project': 'Project',
        'priority': 'Priority',
        'case_applicability': 'Case Applicability',
        'response_time': 'Response Time Category',
        'resolution_time': 'Resolution Time Category',
    }
    for key, val in params.items():
        if val:
            col = col_map.get(key)
            if col:
                filtered = [r for r in filtered if safe_str(r.get(col)) == val]
    return filtered


@case_router.get("/filters")
async def get_case_filters():
    _, rows = load_case_data()

    def unique_sorted(col):
        return sorted(set(safe_str(r.get(col)) for r in rows if safe_str(r.get(col))))

    return {
        "case_types": unique_sorted('Case Type'),
        "statuses": unique_sorted('Status'),
        "case_origins": unique_sorted('Case Origin'),
        "areas": unique_sorted('Area'),
        "sub_areas": unique_sorted('Sub Area'),
        "case_owners": unique_sorted('Case Owner'),
        "hods": unique_sorted('HOD 1'),
        "team_leaders": unique_sorted('Team Leader'),
        "projects": unique_sorted('Project'),
        "priorities": unique_sorted('Priority'),
    }


@case_router.get("/data")
async def get_case_data(
    case_type: Optional[str] = None,
    status: Optional[str] = None,
    case_origin: Optional[str] = None,
    area: Optional[str] = None,
    sub_area: Optional[str] = None,
    case_owner: Optional[str] = None,
    hod: Optional[str] = None,
    team_leader: Optional[str] = None,
    project: Optional[str] = None,
    priority: Optional[str] = None,
    case_applicability: Optional[str] = None,
    response_time: Optional[str] = None,
    resolution_time: Optional[str] = None,
):
    _, rows = load_case_data()
    params = {
        'case_type': case_type, 'status': status, 'case_origin': case_origin,
        'area': area, 'sub_area': sub_area, 'case_owner': case_owner,
        'hod': hod, 'team_leader': team_leader, 'project': project,
        'priority': priority, 'case_applicability': case_applicability,
        'response_time': response_time, 'resolution_time': resolution_time,
    }
    filtered = apply_case_filters(rows, params)

    total = len(filtered)
    open_statuses = {'New', 'In Progress', 'Pending for Clarification', 'Re-Open'}
    closed_statuses = {'Closed', 'Close', 'Resolved'}
    open_count = sum(1 for r in filtered if safe_str(r.get('Status')) in open_statuses)
    closed_count = sum(1 for r in filtered if safe_str(r.get('Status')) in closed_statuses)

    # Status breakdown
    status_counts = {}
    for r in filtered:
        s = safe_str(r.get('Status')) or 'Unknown'
        status_counts[s] = status_counts.get(s, 0) + 1
    status_list = sorted(status_counts.items(), key=lambda x: x[1], reverse=True)

    # Case type breakdown
    type_counts = {}
    for r in filtered:
        t = safe_str(r.get('Case Type')) or 'Unknown'
        type_counts[t] = type_counts.get(t, 0) + 1

    # Case origin breakdown
    origin_counts = {}
    for r in filtered:
        o = safe_str(r.get('Case Origin')) or 'Unknown'
        origin_counts[o] = origin_counts.get(o, 0) + 1
    origin_list = sorted(origin_counts.items(), key=lambda x: x[1], reverse=True)
    origin_pct = [{"name": k, "value": v, "pct": round(v / total * 100, 2) if total > 0 else 0} for k, v in origin_list]

    # Case owner breakdown (top 15)
    owner_counts = {}
    for r in filtered:
        o = safe_str(r.get('Case Owner')) or 'Unknown'
        owner_counts[o] = owner_counts.get(o, 0) + 1
    owner_list = sorted(owner_counts.items(), key=lambda x: x[1], reverse=True)[:15]

    # Area / Sub Area breakdown (top 20)
    area_sub = {}
    for r in filtered:
        a = safe_str(r.get('Area'))
        sa = safe_str(r.get('Sub Area'))
        if a:
            key = f"{a}||{sa}"
            area_sub[key] = area_sub.get(key, 0) + 1
    area_sub_list = sorted(area_sub.items(), key=lambda x: x[1], reverse=True)[:20]
    area_sub_data = [{"area": k.split('||')[0], "sub_area": k.split('||')[1], "count": v} for k, v in area_sub_list]

    # Priority breakdown
    priority_counts = {}
    for r in filtered:
        p = safe_str(r.get('Priority')) or 'Unknown'
        priority_counts[p] = priority_counts.get(p, 0) + 1

    # Project breakdown
    project_counts = {}
    for r in filtered:
        p = safe_str(r.get('Project'))
        if p:
            project_counts[p] = project_counts.get(p, 0) + 1
    project_list = sorted(project_counts.items(), key=lambda x: x[1], reverse=True)

    # Team leader breakdown
    tl_counts = {}
    for r in filtered:
        t = safe_str(r.get('Team Leader'))
        if t:
            tl_counts[t] = tl_counts.get(t, 0) + 1
    tl_list = sorted(tl_counts.items(), key=lambda x: x[1], reverse=True)

    # Response time / Resolution time
    resp_time = {}
    for r in filtered:
        rt = safe_str(r.get('Response Time Category'))
        if rt:
            resp_time[rt] = resp_time.get(rt, 0) + 1

    resol_time = {}
    for r in filtered:
        rt = safe_str(r.get('Resolution Time Category'))
        if rt:
            resol_time[rt] = resol_time.get(rt, 0) + 1

    # Case applicability
    app_counts = {}
    for r in filtered:
        a = safe_str(r.get('Case Applicability'))
        if a:
            app_counts[a] = app_counts.get(a, 0) + 1

    # Escalated count
    escalated = sum(1 for r in filtered if safe_str(r.get('Is Escalated closure')).lower() == 'true')

    # HNI count
    hni = sum(1 for r in filtered if safe_str(r.get('HNI Customer')).lower() == 'true')

    # Active legal
    legal = sum(1 for r in filtered if safe_str(r.get('Active Legal Case')).lower() == 'true')

    return {
        "kpi": {
            "total": total,
            "open": open_count,
            "closed": closed_count,
            "escalated": escalated,
            "hni": hni,
            "legal": legal,
        },
        "status_breakdown": [{"name": k, "value": v} for k, v in status_list],
        "case_types": [{"name": k, "value": v} for k, v in type_counts.items()],
        "case_origins": origin_pct,
        "case_owners": [{"name": k, "value": v} for k, v in owner_list],
        "area_sub_area": area_sub_data,
        "priorities": [{"name": k, "value": v} for k, v in priority_counts.items()],
        "projects": [{"name": k, "value": v} for k, v in project_list],
        "team_leaders": [{"name": k, "value": v} for k, v in tl_list],
        "response_time": [{"name": k, "value": v} for k, v in resp_time.items()],
        "resolution_time": [{"name": k, "value": v} for k, v in resol_time.items()],
        "applicability": [{"name": k, "value": v} for k, v in app_counts.items()],
    }


@case_router.get("/table")
async def get_case_table(
    case_type: Optional[str] = None,
    status: Optional[str] = None,
    case_origin: Optional[str] = None,
    area: Optional[str] = None,
    sub_area: Optional[str] = None,
    case_owner: Optional[str] = None,
    hod: Optional[str] = None,
    team_leader: Optional[str] = None,
    project: Optional[str] = None,
    priority: Optional[str] = None,
    case_applicability: Optional[str] = None,
    response_time: Optional[str] = None,
    resolution_time: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    search: Optional[str] = None,
):
    _, rows = load_case_data()
    params = {
        'case_type': case_type, 'status': status, 'case_origin': case_origin,
        'area': area, 'sub_area': sub_area, 'case_owner': case_owner,
        'hod': hod, 'team_leader': team_leader, 'project': project,
        'priority': priority, 'case_applicability': case_applicability,
        'response_time': response_time, 'resolution_time': resolution_time,
    }
    filtered = apply_case_filters(rows, params)

    if search:
        s = search.lower()
        filtered = [r for r in filtered if
                    s in safe_str(r.get('Case Number')).lower() or
                    s in safe_str(r.get('Account Name')).lower() or
                    s in safe_str(r.get('Subject')).lower() or
                    s in safe_str(r.get('Case Owner')).lower()]

    total = len(filtered)
    start = (page - 1) * page_size
    end = start + page_size
    page_rows = filtered[start:end]

    table_data = []
    for r in page_rows:
        table_data.append({
            "case_number": safe_str(r.get('Case Number')),
            "account_name": safe_str(r.get('Account Name')),
            "case_owner": safe_str(r.get('Case Owner')),
            "hod": safe_str(r.get('HOD 1')),
            "team_leader": safe_str(r.get('Team Leader')),
            "status": safe_str(r.get('Status')),
            "case_type": safe_str(r.get('Case Type')),
            "case_origin": safe_str(r.get('Case Origin')),
            "area": safe_str(r.get('Area')),
            "sub_area": safe_str(r.get('Sub Area')),
            "priority": safe_str(r.get('Priority')),
            "project": safe_str(r.get('Project')),
            "date_opened": safe_str(r.get('Date/Time Opened')),
            "closed_date": safe_str(r.get('Closed Date')),
            "subject": safe_str(r.get('Subject'))[:80],
            "response_time": safe_str(r.get('Response Time Category')),
            "resolution_time": safe_str(r.get('Resolution Time Category')),
            "case_applicability": safe_str(r.get('Case Applicability')),
        })

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
        "rows": table_data,
    }


@case_router.post("/upload")
async def upload_case_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files accepted")
    if CASE_FILE.exists():
        shutil.copy2(CASE_FILE, DATA_DIR / 'CaseManagement_backup.xlsx')
    contents = await file.read()
    with open(CASE_FILE, 'wb') as f:
        f.write(contents)
    try:
        wb = openpyxl.load_workbook(CASE_FILE, read_only=True)
        ws = wb['Sheet1']
        count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        _cache["rows"] = None  # Invalidate cache
        _cache["mtime"] = 0
        return {"message": "File uploaded successfully", "rows_count": count}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid file: {str(e)}")


@case_router.get("/download")
async def download_case_file():
    if not CASE_FILE.exists():
        raise HTTPException(status_code=404, detail="No case file found")
    return FileResponse(
        str(CASE_FILE),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename='CaseManagement.xlsx'
    )
