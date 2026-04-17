from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from typing import Optional
import openpyxl
import shutil

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

app = FastAPI()
api_router = APIRouter(prefix="/api")

DATA_DIR = ROOT_DIR / 'data'
DATA_DIR.mkdir(exist_ok=True)
ZALR_FILE = DATA_DIR / 'ZALR.xlsx'

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def ensure_seed_data():
    if not ZALR_FILE.exists():
        from seed_data import generate_seed_excel
        generate_seed_excel()


_zalr_cache = {"rows": None, "budgets": None, "mtime": 0}


def safe_num(v):
    """Convert value to float, return 0 for None, formulas, or invalid."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        # Skip Excel formula strings
        if v.strip().startswith('='):
            return 0
        try:
            return float(str(v).replace(',', '').strip())
        except (ValueError, TypeError):
            return 0
    return 0


def read_zalr_excel():
    ensure_seed_data()
    mtime = ZALR_FILE.stat().st_mtime
    if _zalr_cache["rows"] is not None and _zalr_cache["mtime"] == mtime:
        return _zalr_cache["rows"], _zalr_cache["budgets"]

    # Use data_only=True to get calculated values instead of formulas
    wb = openpyxl.load_workbook(ZALR_FILE, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    raw_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                d[headers[i]] = val
        raw_rows.append(d)
    wb.close()

    rows = []
    budgets = {}

    import datetime as dt_module

    for d in raw_rows:
        # Parse document date
        doc_date = d.get('Document Date')
        yr, mo = None, None
        if isinstance(doc_date, (dt_module.datetime, dt_module.date)):
            yr = doc_date.year
            mo = doc_date.month
        elif isinstance(doc_date, str) and doc_date:
            try:
                parsed = dt_module.datetime.strptime(doc_date.split(' ')[0], '%Y-%m-%d')
                yr, mo = parsed.year, parsed.month
            except Exception:
                pass

        # Your Excel columns — map them directly
        ordered_gst      = safe_num(d.get('Ordered with GST'))
        delivered_gst    = safe_num(d.get('Delivered with GST'))
        invoiced_gst     = safe_num(d.get('Invoiced with GST'))
        still_del_gst    = safe_num(d.get('Still to be Delivered with GST'))
        still_inv_gst    = safe_num(d.get('Still to be Invoiced with GST'))
        budget           = safe_num(d.get('Budget'))
        actual           = safe_num(d.get('Actual'))
        assigned         = safe_num(d.get('Assigned'))
        commitment       = safe_num(d.get('Commitment'))
        available        = safe_num(d.get('Available'))

        # If Still to be Invoiced with GST was a formula, compute it
        if still_inv_gst == 0 and invoiced_gst > 0:
            still_inv_gst = max(0, ordered_gst - invoiced_gst)

        wbs_key  = str(d.get('WBS', '') or '')
        plant    = str(d.get('Plant', '') or '')
        proj_tag = str(d.get('PROJECT/NON-PROJECT', '') or '')

        # Normalise Project/Non-Project labels
        if 'non' in proj_tag.lower():
            proj_type = 'Non-Project'
        elif 'proj' in proj_tag.lower():
            proj_type = 'Project'
        else:
            proj_type = proj_tag

        mapped = {
            'WBS Element':          wbs_key,
            'WBS Description':      str(d.get('WBS Description', '') or ''),
            'Plant':                plant,
            'Plant Name':           str(d.get('Plant Name', '') or ''),
            'Purchasing Document':  d.get('Purchasing Document'),
            'Project/Non-Project':  proj_type,
            'Work Type':            str(d.get('WORK TYPE', '') or ''),
            'Vendor Name':          str(d.get('Vendor Name', '') or ''),
            'Short Text':           str(d.get('Short Text', '') or ''),
            'Document Type':        str(d.get('Document Type', '') or ''),
            'Document Date':        doc_date,
            'Year':                 yr,
            'Month':                mo,
            'Budget':               budget,
            'Actual':               actual,
            'Assigned':             assigned,
            'Commitment':           commitment,
            'Available':            available,
            'Ordered with GST':     ordered_gst,
            'Delivered with GST':   delivered_gst,
            'Invoiced with GST':    invoiced_gst,
            'Still to Deliver GST': still_del_gst,
            'Still to Invoice':     still_inv_gst,
            'Ordered Value':        safe_num(d.get('Ordered Value')),
            'Net Price':            safe_num(d.get('Net Price')),
            'Currency':             str(d.get('Currency', '') or ''),
        }
        rows.append(mapped)

        # Build budget map per WBS (use max value seen)
        if wbs_key:
            if budget and (wbs_key not in budgets or budget > budgets.get(wbs_key, 0)):
                budgets[wbs_key] = budget

    _zalr_cache["rows"]    = rows
    _zalr_cache["budgets"] = budgets
    _zalr_cache["mtime"]   = mtime
    logger.info(f"Loaded {len(rows)} rows, {len(budgets)} WBS budgets")
    return rows, budgets


PLANT_NAMES = {
    "1000": "SWD One DXP-Ph1 Res.", "1010": "SmartWorld Developer",
    "1011": "SWD OneDXP-Ph2 Stret", "1012": "One DXP-Commercial",
    "1013": "SWD One DXP-Ph2 Res.", "1014": "One DXP Phase-5",
    "1015": "One DXP Sales Gallery", "1070": "Riverday Resi.-69",
    "1071": "Riverday Retail-69",   "1072": "Trump Tower",
    "1073": "IFC 11th Floor",       "1074": "Lead Apartment-Trump",
    "1090": "Anuvridhi Head Offc.", "2000": "Smartworld Heights",
    "2010": "ETSY Developer P.Ltd", "2011": "SWD Central Office",
    "2012": "Sales Gallery",        "2070": "Manesar M11-HO",
    "2071": "Manesar M11-Project",  "2072": "Manesar M11-Commer.",
    "2080": "Topshelf Builders Re", "3010": "Glorii Education-HO",
    "3070": "Sector 98 Noida HO",  "3071": "SWD Residencies",
    "3072": "SWD Le Courtyard",     "3073": "Smartworld Suites",
    "3074": "Sector 98 Noida SGE",
}


def process_data(rows, budgets, plant=None, wbs=None, po=None,
                 proj_type=None, year=None, month=None, work_type=None):
    filtered = rows
    if plant:
        ps = set(plant.split(','))
        filtered = [r for r in filtered if r.get('Plant', '') in ps]
    if wbs:
        ws = set(wbs.split(','))
        filtered = [r for r in filtered if r.get('WBS Element', '') in ws]
    if po:
        pos = set(po.split(','))
        filtered = [r for r in filtered if str(r.get('Purchasing Document', '')) in pos]
    if proj_type:
        filtered = [r for r in filtered if r.get('Project/Non-Project', '') == proj_type]
    if year:
        ys = set(year.split(','))
        filtered = [r for r in filtered if str(r.get('Year', '')) in ys]
    if month:
        ms = set(month.split(','))
        filtered = [r for r in filtered if str(r.get('Month', '')) in ms]
    if work_type:
        wts = set(work_type.split(','))
        filtered = [r for r in filtered if r.get('Work Type', '').strip() in wts]
    return filtered


def aggregate(rows_list):
    return {
        "ordered":          sum(r.get('Ordered with GST', 0) or 0 for r in rows_list),
        "delivered":        sum(r.get('Delivered with GST', 0) or 0 for r in rows_list),
        "invoiced":         sum(r.get('Invoiced with GST', 0) or 0 for r in rows_list),
        "still_to_invoice": sum(r.get('Still to Invoice', 0) or 0 for r in rows_list),
        "still_to_deliver": sum(r.get('Still to Deliver GST', 0) or 0 for r in rows_list),
        "actual":           sum(r.get('Actual', 0) or 0 for r in rows_list),
        "commitment":       sum(r.get('Commitment', 0) or 0 for r in rows_list),
    }


def get_unique_wbs(rows_list):
    return list(set(r.get('WBS Element', '') for r in rows_list if r.get('WBS Element')))


@api_router.get("/")
async def root():
    return {"message": "SmartWorld Analytics API"}


@api_router.get("/data")
async def get_data(
    plant: Optional[str] = None, wbs: Optional[str] = None,
    po: Optional[str] = None, proj_type: Optional[str] = None,
    year: Optional[str] = None, month: Optional[str] = None,
    work_type: Optional[str] = None
):
    rows, budgets = read_zalr_excel()
    filtered  = process_data(rows, budgets, plant, wbs, po, proj_type, year, month, work_type)
    proj_rows = [r for r in filtered if r.get('Project/Non-Project') == 'Project']
    non_rows  = [r for r in filtered if r.get('Project/Non-Project') == 'Non-Project']

    tot = aggregate(filtered)
    tp  = aggregate(proj_rows)
    tn  = aggregate(non_rows)

    wbs_all  = get_unique_wbs(filtered)
    wbs_proj = get_unique_wbs(proj_rows)
    wbs_non  = get_unique_wbs(non_rows)
    po_all   = list(set(str(r.get('Purchasing Document', '')) for r in filtered if r.get('Purchasing Document')))

    bud_all  = sum(budgets.get(w, 0) for w in wbs_all)
    bud_proj = sum(budgets.get(w, 0) for w in wbs_proj)
    bud_non  = sum(budgets.get(w, 0) for w in wbs_non)

    # Monthly trend
    monthly = {}
    for r in filtered:
        yr, mo = r.get('Year'), r.get('Month')
        if not yr or not mo:
            continue
        key = f"{yr}-{int(mo):02d}"
        if key not in monthly:
            monthly[key] = {"year": yr, "month": int(mo), "ord_proj": 0, "del_proj": 0, "ord_non": 0, "del_non": 0}
        is_proj = r.get('Project/Non-Project') == 'Project'
        if is_proj:
            monthly[key]["ord_proj"] += r.get('Ordered with GST', 0) or 0
            monthly[key]["del_proj"] += r.get('Delivered with GST', 0) or 0
        else:
            monthly[key]["ord_non"] += r.get('Ordered with GST', 0) or 0
            monthly[key]["del_non"] += r.get('Delivered with GST', 0) or 0
    monthly_sorted = sorted(monthly.values(), key=lambda x: x['year'] * 100 + x['month'])

    # Plant-wise
    plant_data = {}
    for r in filtered:
        pl = str(r.get('Plant', ''))
        if pl not in plant_data:
            plant_data[pl] = {"plant": pl, "plant_name": PLANT_NAMES.get(pl, ''), "ordered": 0, "delivered": 0, "still_to_deliver": 0}
        plant_data[pl]["ordered"]        += r.get('Ordered with GST', 0) or 0
        plant_data[pl]["delivered"]      += r.get('Delivered with GST', 0) or 0
        plant_data[pl]["still_to_deliver"] += r.get('Still to Deliver GST', 0) or 0
    plant_sorted = sorted(plant_data.values(), key=lambda x: x['ordered'], reverse=True)

    # WBS budget pie top 10
    wbs_bud_list = [{"wbs": w, "budget": budgets.get(w, 0), "is_project": w.startswith('RE/')} for w in wbs_all if budgets.get(w, 0) > 0]
    wbs_bud_list.sort(key=lambda x: x['budget'], reverse=True)

    # Yearly stacked
    yearly = {}
    for r in filtered:
        yr = str(r.get('Year', '') or '')
        if not yr or yr == 'None':
            continue
        if yr not in yearly:
            yearly[yr] = {"year": yr, "ordered": 0, "delivered": 0, "invoiced": 0, "still_to_deliver": 0}
        yearly[yr]["ordered"]        += r.get('Ordered with GST', 0) or 0
        yearly[yr]["delivered"]      += r.get('Delivered with GST', 0) or 0
        yearly[yr]["invoiced"]       += r.get('Invoiced with GST', 0) or 0
        yearly[yr]["still_to_deliver"] += r.get('Still to Deliver GST', 0) or 0
    yearly_sorted = sorted(yearly.values(), key=lambda x: x['year'])

    # WBS ordered top 10
    wbs_ord_map = {}
    for r in filtered:
        w = r.get('WBS Element', '')
        if not w:
            continue
        if w not in wbs_ord_map:
            wbs_ord_map[w] = {"wbs": w, "ordered": 0, "delivered": 0, "still_to_deliver": 0, "is_project": r.get('Project/Non-Project') == 'Project'}
        wbs_ord_map[w]["ordered"]          += r.get('Ordered with GST', 0) or 0
        wbs_ord_map[w]["delivered"]        += r.get('Delivered with GST', 0) or 0
        wbs_ord_map[w]["still_to_deliver"] += r.get('Still to Deliver GST', 0) or 0
    wbs_ord_sorted = sorted(wbs_ord_map.values(), key=lambda x: x['ordered'], reverse=True)

    # Plant utilization
    plant_util = []
    for pd in plant_sorted[:8]:
        pl = pd['plant']
        wbs_in_plant = set(r.get('WBS Element') for r in filtered if str(r.get('Plant', '')) == pl and r.get('WBS Element'))
        bud = sum(budgets.get(w, 0) for w in wbs_in_plant)
        pct = min(100, round(pd['ordered'] / bud * 100)) if bud > 0 else 0
        plant_util.append({"plant": pl, "plant_name": PLANT_NAMES.get(pl, ''), "ordered": pd['ordered'], "budget": bud, "utilization": pct})

    # WBS table top 25
    wbs_table = []
    desc_map = {}
    proj_map = {}
    for r in filtered:
        w = r.get('WBS Element', '')
        if w and w not in desc_map:
            desc_map[w] = r.get('WBS Description', '')
            proj_map[w] = r.get('Project/Non-Project', '')

    for entry in wbs_ord_sorted[:25]:
        w   = entry['wbs']
        bud = budgets.get(w, 0)
        still_inv = sum((r.get('Still to Invoice', 0) or 0) for r in filtered if r.get('WBS Element') == w)
        invoiced  = sum((r.get('Invoiced with GST', 0) or 0) for r in filtered if r.get('WBS Element') == w)
        wbs_table.append({
            "type": proj_map.get(w, ''), "wbs": w, "description": desc_map.get(w, ''),
            "budget": bud, "ordered": entry['ordered'],
            "delivered": entry['delivered'], "invoiced": invoiced,
            "still_to_deliver": entry['still_to_deliver'], "still_to_invoice": still_inv
        })

    # Work type breakdown (NEW - from your WORK TYPE column)
    work_type_data = {}
    for r in filtered:
        wt = str(r.get('Work Type', '') or '').strip()
        if not wt:
            wt = 'Unclassified'
        if wt not in work_type_data:
            work_type_data[wt] = {"work_type": wt, "ordered": 0, "delivered": 0, "count": 0}
        work_type_data[wt]["ordered"]   += r.get('Ordered with GST', 0) or 0
        work_type_data[wt]["delivered"] += r.get('Delivered with GST', 0) or 0
        work_type_data[wt]["count"]     += 1
    work_type_sorted = sorted(work_type_data.values(), key=lambda x: x['ordered'], reverse=True)

    return {
        "kpi": {
            "wbs_count": len(wbs_all), "wbs_proj": len(wbs_proj), "wbs_non": len(wbs_non),
            "budget": bud_all, "budget_proj": bud_proj, "budget_non": bud_non,
            "po_count": len(po_all),
            "po_proj": len(set(str(r.get('Purchasing Document', '')) for r in proj_rows if r.get('Purchasing Document'))),
            "po_non":  len(set(str(r.get('Purchasing Document', '')) for r in non_rows if r.get('Purchasing Document'))),
            "total": tot, "proj": tp, "non_proj": tn
        },
        "row_count": len(filtered),
        "proj_count": len(proj_rows),
        "non_count":  len(non_rows),
        "monthly_trend":    monthly_sorted,
        "plant_data":       plant_sorted,
        "wbs_budget_top":   wbs_bud_list[:10],
        "yearly_data":      yearly_sorted,
        "wbs_ordered_top":  wbs_ord_sorted[:10],
        "plant_utilization": plant_util,
        "wbs_table":        wbs_table,
        "work_type_data":   work_type_sorted,
    }


@api_router.get("/filters")
async def get_filters():
    rows, budgets = read_zalr_excel()

    plant_codes = sorted(set(str(r.get('Plant', '')) for r in rows if r.get('Plant')))
    plants_with_names = [{"value": p, "label": f"{p} — {PLANT_NAMES.get(p, '')}" if PLANT_NAMES.get(p) else p} for p in plant_codes]

    wbs_list = sorted(set(r.get('WBS Element', '') for r in rows if r.get('WBS Element')))
    desc_map = {}
    for r in rows:
        w = r.get('WBS Element', '')
        if w and w not in desc_map:
            desc_map[w] = r.get('WBS Description', '')
    wbs_with_desc = [{"value": w, "label": f"{w} — {(desc_map.get(w, '') or '')[:30]}"} for w in wbs_list]

    pos   = sorted(set(str(r.get('Purchasing Document', '')) for r in rows if r.get('Purchasing Document')))[:600]
    years = sorted(set(str(int(r.get('Year'))) for r in rows if r.get('Year') is not None))
    work_types = sorted(set(str(r.get('Work Type', '') or '').strip() for r in rows if r.get('Work Type') and str(r.get('Work Type','')).strip()))

    return {
        "plants":               plants_with_names,
        "wbs_elements":         wbs_with_desc,
        "purchasing_documents": pos,
        "years":                years,
        "work_types":           work_types,
    }


@api_router.get("/stats")
async def get_stats():
    rows, budgets = read_zalr_excel()
    total_budget = sum(budgets.values())
    return {
        "wbs_elements":   len(set(r.get('WBS Element', '') for r in rows if r.get('WBS Element'))),
        "purchase_orders": len(set(str(r.get('Purchasing Document', '')) for r in rows if r.get('Purchasing Document'))),
        "total_budget":   total_budget,
        "plants":         len(set(str(r.get('Plant', '')) for r in rows if r.get('Plant'))),
        "report_modules": 8,
        "total_rows":     len(rows),
    }


@api_router.get("/modules")
async def get_modules():
    return [
        {"id": "zalr",            "title": "ZALR Cost Dashboard",       "description": "Procurement analytics — Budget, Ordered, Delivered, Invoiced, Still to Deliver with GST. Full plant-wise and WBS-wise breakdown.", "icon": "wallet",      "status": "live"},
        {"id": "project-progress","title": "Project Progress Report",   "description": "Track milestone completion, schedule variance, and project health.",                                                               "icon": "bar-chart-3", "status": "coming_soon"},
        {"id": "vendor",          "title": "Vendor Performance",        "description": "Supplier scorecard with delivery timelines, quality metrics, and spend analysis.",                                                "icon": "factory",     "status": "coming_soon"},
        {"id": "budget-variance", "title": "Budget Variance Report",    "description": "Actuals vs budget comparison with variance flags and departmental drill-down.",                                                   "icon": "trending-up", "status": "coming_soon"},
        {"id": "hr",              "title": "HR Analytics",              "description": "Headcount, payroll cost tracking, and workforce distribution.",                                                                    "icon": "users",       "status": "coming_soon"},
        {"id": "sales",           "title": "Sales Dashboard",           "description": "Unit booking status, collection tracking, payment summary and ageing report.",                                                    "icon": "building-2",  "status": "live"},
        {"id": "cases",           "title": "Case Management",           "description": "Ticket status, case origin, owner workload, TAT monitoring and detailed case records.",                                           "icon": "search",      "status": "live"},
        {"id": "finance",         "title": "Finance MIS Report",        "description": "Receivables, payables, cash flow projections and consolidated P&L.",                                                              "icon": "receipt",     "status": "coming_soon"},
    ]


@api_router.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, .csv files are accepted")

    contents = await file.read()
    temp_path = DATA_DIR / 'temp_upload.xlsx'
    with open(temp_path, 'wb') as f:
        f.write(contents)

    try:
        wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        first_sheet = wb[sheet_names[0]]
        headers = [cell.value for cell in next(first_sheet.iter_rows(min_row=1, max_row=1))]
        has_wbs = 'WBS' in headers or 'WBS Element' in headers
        wb.close()
        if not has_wbs:
            temp_path.unlink()
            raise HTTPException(status_code=400, detail="Missing required 'WBS' column")
    except HTTPException:
        raise
    except Exception as e:
        if temp_path.exists():
            temp_path.unlink()
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    # Backup old file
    if ZALR_FILE.exists():
        shutil.copy2(ZALR_FILE, DATA_DIR / 'ZALR_backup.xlsx')

    shutil.move(str(temp_path), str(ZALR_FILE))

    # Invalidate cache
    _zalr_cache["rows"]  = None
    _zalr_cache["mtime"] = 0

    rows, budgets = read_zalr_excel()
    return {"message": "File uploaded successfully", "rows_count": len(rows)}


@api_router.get("/download")
async def download_current():
    from fastapi.responses import FileResponse
    ensure_seed_data()
    return FileResponse(str(ZALR_FILE), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename='ZALR.xlsx')


from sales_api import sales_router
from case_api import case_router
app.include_router(api_router)
app.include_router(sales_router)
app.include_router(case_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup():
    ensure_seed_data()
    from sales_api import ensure_sales_data
    ensure_sales_data()
    logger.info("SmartWorld Analytics API started")
